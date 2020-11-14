import openpyxl, os, csv
from openpyxl import Workbook
folderName = "oops"

def getFolderList(): #creates a list of all folders in Results
    global superFolder
    global folderList
    superFolder = 'C:\\Users\\localadmin\\Results\\'
    os.chdir(superFolder)
    folderList = []
    for resultFolder in os.listdir(os.getcwd()):
        if os.path.isdir(resultFolder):
            folderList.append(resultFolder)


#def assignmentList(folderName): #creates a list with the names of all text files containing data in each folder
#    global assignmentList
#    print("folderName is " + folderName)
#    assignmentList = [os.path.splitext(assignment)[0] for assignment in os.listdir(os.getcwd()) if assignment.endswith('.txt')]


def numberFinder(string, place): #checks if a certain word in a string is a number
    count = 0
    for word in string.split("_"):
        if word[0].isdigit():
            if count == place:
                return word
                break
            count += 1

def makeExcel(folderName): #creates the excel spreadsheet
    print("Current Folder is " + folderName)
    os.chdir(folderName)
    global wb
    global sheet
    wb = Workbook()
    wb.create_sheet("Label")
    sheet = wb.active

    
    

def makeHeader(): #adds the first header row to the workbook containing fields and trials
    headers = ["Box", "Date/Time","Assignment","Gender","Concentration","Fault Out"]
    for count, header in enumerate(headers, 1):
        sheet.cell(1,count).value = header

    for i in range(1,31): #trials
        sheet.cell(1, len(headers) + i).value = "T" + str(i)

def excelFiller(folderName): #Fills in excel fields
    assignmentList = ['acceptSide', 'acceptTime', 'numSideSwaps', 'rejectTime', 'seekSideSwaps', 'shockedTime', 'shockModeTime', 'timeToAccept']
    makeExcel(folderName)
    makeHeader()
    print("folderName is " + folderName)
    #assignmentList(folderName)
    sheetRow = 2
    global assignmentLabels
    global assignmentLabelCounter
    #assignmentLabels = ["trialNumber", "seekSideSwaps","timetoAccept","acceptSide","shockModeTime","acceptTime","numSideSwaps"]
    assignmentLabels = ["trialNumber","shockedTime","seekSideSwaps","timetoAccept","acceptSide","shockModeTime","acceptTime","numSideSwaps"]
    assignmentLabelCounter = 0
    for place, assign in enumerate(assignmentList): #for each assignment, put in one row per line of text in the txt files
        try:
            file = open(assign + ".txt") 
            reader = csv.reader(file)
            for row in reader: #csv reader outputs one object for each line of text in a txt file
                if row[0] == "FAULT OUT": #mark if fault out
                    sheet.cell(sheetRow, 6).value= "Yes"
                    del row[0]
                gender = row[3] #collect data fields
                concentration = row[1]
                boxNum = numberFinder(row[2], 0)
                data = []
                for i in range(4, len(row)): 
                    data.append(row[i])
                sheet.cell(sheetRow, 1).value = boxNum 
                sheet.cell(sheetRow, 2).value = numberFinder(row[2], 1)
                sheet.cell(sheetRow, 3).value = assignmentLabels[place]
                #sheet.cell(sheetRow, 3).value = assignmentLabels[assignmentLabelCounter]
                sheet.cell(sheetRow, 4).value = gender
                sheet.cell(sheetRow, 5).value = concentration
                for i in range(len(data)):
                    sheet.cell(sheetRow, i + 7).value = data[i]
                
                sheetRow += 1
                #print("sheetRow is " + str(sheetRow))
            #assignmentLabelCounter += 1
            #print("assignmentLabelCounter is " + str(place))
            #print("Label is " + assignmentLabels[place])
            file.seek(0)

        except FileNotFoundError:
            print("There are some text files missing here!")
            
    #print("ExcelFiller has run")
        
    folderPath = 'C:\\Users\\localadmin\\Results\\' + folderName
    #wb.save(os.path.join(folderPath, folderName + ".xlsx"))
    wb.save(folderName + ".xlsx")
    os.chdir(superFolder)

getFolderList()
print("folderList is " + str(folderList))
for counting, folder in enumerate(folderList):
    excelFiller(folder)


print("Job Done!")
    


